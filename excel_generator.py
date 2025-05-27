import openpyxl
import json
import sys
import os

def create_input_demand_excel(filepath):
    workbook = openpyxl.Workbook()
    
    # Main sheet
    sheet_main = workbook.active
    sheet_main.title = "main"
    
    main_settings_data = [
        ["~Settings", None],
        ["Parameters", "Inputs"],
        ["Start_Year", 2015],
        ["End_Year", 2025],
        ["Econometric_Parameters", "Yes"]
    ]
    for r, row_data in enumerate(main_settings_data, start=1):
        for c, cell_data in enumerate(row_data, start=1):
            sheet_main.cell(row=r, column=c, value=cell_data)

    main_consumption_sectors_data = [
        ["~Consumption_Sectors"],
        ["Sector_Name"],
        ["Residential"],
        ["Commercial"]
    ]
    current_row = sheet_main.max_row + 2
    for r, row_data in enumerate(main_consumption_sectors_data, start=current_row):
        for c, cell_data in enumerate(row_data, start=1):
            sheet_main.cell(row=r, column=c, value=cell_data)
            
    main_econometric_parameters_data = [
        ["~Econometric_Parameters", None],
        ["Residential", "Commercial"],
        ["GDP", "Population"],
        ["Population", None]
    ]
    current_row = sheet_main.max_row + 2
    for r, row_data in enumerate(main_econometric_parameters_data, start=current_row):
        for c, cell_data in enumerate(row_data, start=1):
            sheet_main.cell(row=r, column=c, value=cell_data)

    # Economic_Indicators sheet
    sheet_econ = workbook.create_sheet("Economic_Indicators")
    econ_headers = ["Year", "GDP", "Population"]
    sheet_econ.append(econ_headers)
    for i in range(11): # 2015 to 2025
        year = 2015 + i
        gdp = 100 + (i * 5)
        population = 10 + i
        sheet_econ.append([year, gdp, population])

    # Residential sheet
    sheet_res = workbook.create_sheet("Residential")
    res_headers = ["Year", "Electricity", "SomeOtherData"]
    sheet_res.append(res_headers)
    for i in range(11):
        year = 2015 + i
        electricity = 50 + (i * 2)
        some_other_data = 5 + int(i * 0.5)
        sheet_res.append([year, electricity, some_other_data])

    # Commercial sheet
    sheet_com = workbook.create_sheet("Commercial")
    com_headers = ["Year", "Electricity", "AnotherValue"]
    sheet_com.append(com_headers)
    for i in range(11):
        year = 2015 + i
        electricity = 30 + int(i * 1.5) # 30, 31, 33, 34, 36, 37, 39, 40, 42, 43, 45
        another_value = 3 + int(i * 0.4) # 3, 3, 3, 4, 4, 5, 5, 5, 6, 6, 7
        sheet_com.append([year, electricity, another_value])
        
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    workbook.save(filepath)
    print(f"Created input demand Excel: {filepath}")

def create_sector_result_excel(filepath, sector_name):
    workbook = openpyxl.Workbook()
    sheet_results = workbook.active
    sheet_results.title = "Results"
    
    if sector_name == "Residential":
        headers = ["Year", "Historical", "ModelA", "ModelB"]
        sheet_results.append(headers)
        for i in range(11): # 2015 to 2025
            year = 2015 + i
            historical = 50 + (i * 2)
            model_a = 51 + int(i * 2.1) # 51, 53, 55, 57, 59, 61, 63, 65, 67, 69, 72
            model_b = 52 + int(i * 2.1) # 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 73
            sheet_results.append([year, historical, model_a, model_b])
    elif sector_name == "Commercial":
        headers = ["Year", "Historical", "ModelX"]
        sheet_results.append(headers)
        for i in range(11): # 2015 to 2025
            year = 2015 + i
            historical = 30 + int(i * 1.5)
            model_x = 31 + int(i * 1.5) # 31, 32, 34, 35, 37, 38, 40, 41, 43, 44, 46
            sheet_results.append([year, historical, model_x])
            
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    workbook.save(filepath)
    print(f"Created sector result Excel: {filepath}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_generator.py <output_filepath> <file_type> [sector_name_if_result]")
        sys.exit(1)
        
    output_filepath = sys.argv[1]
    file_type = sys.argv[2]
    
    if file_type == "input_demand":
        create_input_demand_excel(output_filepath)
    elif file_type == "sector_result":
        if len(sys.argv) < 4:
            print("Error: Sector name required for sector_result file type.")
            sys.exit(1)
        sector_name = sys.argv[3]
        create_sector_result_excel(output_filepath, sector_name)
    else:
        print(f"Error: Unknown file_type '{file_type}'")
        sys.exit(1)
